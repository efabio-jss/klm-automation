import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import win32com.client as win32

# === CONFIGURATION ===
base_path = "Admin_Operations_Automation"
master_path = os.path.join(base_path, "Master.xlsx")
template_path = os.path.join(base_path, "Template_Mapa_KM.xlsx")

mes = datetime.now().strftime("%B")
ano = datetime.now().year

output_folder = os.path.join(base_path, "Mapas_Gerados", f"{mes}_{ano}")
os.makedirs(output_folder, exist_ok=True)

# === Read Master Data ===
master_data = pd.read_excel(master_path, sheet_name=None)

for sheet_name, df in master_data.items():
    company_name = sheet_name.strip()
    df.columns = [str(col).strip() for col in df.columns]
    if "KLM" not in df.columns or df["KLM"].isnull().all():
        continue
    df = df[df["KLM"] > 0]
    if df.empty:
        continue

    wb = load_workbook(template_path)
    if "{{empresa}}" not in wb.sheetnames:
        continue
    template_sheet = wb["{{empresa}}"]

    for employee_name, employee_data in df.groupby("Nome"):
        if not employee_name or employee_name.lower().strip() == "colaborador":
            continue

        sheet = wb.copy_worksheet(template_sheet)
        sheet.title = employee_name[:31]
        total_km = employee_data["KLM"].sum()
        first = employee_data.iloc[0]

        sheet["A3"] = company_name
        sheet["A4"] = first.get("Nome Empresa {{empresa}}", "")
        sheet["B5"] = first.get("Morada Empresa {{moradaempresa}}", "")
        sheet["C6"] = first.get("NIF Empresa {{nifempresa}}", "")
        sheet["E5"] = first.get("Mes {{mes}}", mes)
        sheet["F5"] = first.get("Ano {{ano}}", ano)
        sheet["F40"] = total_km
        sheet["B45"] = employee_name
        sheet["B46"] = first.get("Morada {{morada}}", "")
        sheet["E45"] = first.get("NIF Colaborador {{nifcolaborador}}", "")
        sheet["E46"] = first.get("Viatura {{viatura}}", "")

        used_rows = []
        for _, row in employee_data.iterrows():
            try:
                dia = int(float(row.get("Dia  {{dia}}", 0)))
                if 1 <= dia <= 30:
                    r = 8 + dia
                    sheet[f"A{r}"] = dia
                    sheet[f"B{r}"] = row.get("Local de Origem {{localdeorigem}}", "")
                    sheet[f"C{r}"] = row.get("Local de Destino {{localdedestino}}", "")
                    sheet[f"D{r}"] = row.get("Tempo Objectivo {{tempoobjectivo}}", "")
                    sheet[f"F{r}"] = row.get("KLM", "")
                    used_rows.append(r)
            except:
                continue

        for r in range(9, 39):
            sheet[f"A{r}"] = r - 8
            if r not in used_rows:
                for col in ["B", "C", "D", "F"]:
                    sheet[f"{col}{r}"] = ""

    wb.remove(template_sheet)
    output_path = os.path.join(output_folder, f"{mes}_{company_name}.xlsx")
    wb.save(output_path)
    print(f"✅ Created: {output_path}")

    # === KPI Update ===
    kpi_path = os.path.join(base_path, "KPIs")
    os.makedirs(kpi_path, exist_ok=True)
    kpi_file = os.path.join(kpi_path, "KPI.xlsx")

    wb_read = load_workbook(output_path, data_only=True)
    kpi_entries = []

    for sheet_name in wb_read.sheetnames:
        sheet = wb_read[sheet_name]
        try:
            kms = sheet["F40"].value
            rate = sheet["D42"].value
            amount = round(float(kms) * float(rate), 2) if isinstance(kms, (int, float)) and isinstance(rate, (int, float)) else ""
            kpi_entries.append({
                "Month": mes,
                "Year": ano,
                "Company": company_name,
                "Employee": sheet_name,
                "KM": kms,
                "Value (€)": amount
            })
        except:
            continue

    if os.path.exists(kpi_file):
        df_kpi = pd.read_excel(kpi_file)
        df_kpi = pd.concat([df_kpi, pd.DataFrame(kpi_entries)], ignore_index=True)
    else:
        df_kpi = pd.DataFrame(kpi_entries)

    df_kpi.to_excel(kpi_file, index=False)

    # === Charts ===
    wb_kpi = load_workbook(kpi_file)
    if "Charts" in wb_kpi.sheetnames:
        del wb_kpi["Charts"]
    ws_chart = wb_kpi.create_sheet("Charts")

    grouped = df_kpi.groupby(["Company", "Employee"], as_index=False).agg({
        "KM": "sum",
        "Value (€)": "sum"
    })

    start_row = 1
    for company in grouped["Company"].unique():
        subset = grouped[grouped["Company"] == company]
        ws_chart.cell(row=start_row, column=1, value=f"Company: {company}")
        ws_chart.append(["Employee", "KM", "Value (€)"])
        for _, r in subset.iterrows():
            ws_chart.append([r["Employee"], r["KM"], r["Value (€)"]])

        chart = BarChart()
        chart.title = f"{company} - Totals per Employee"
        chart.y_axis.title = "KM / Value (€)"
        chart.height = 7
        chart.width = 15

        data_start = start_row + 1
        data_end = data_start + len(subset)
        data = Reference(ws_chart, min_col=2, max_col=3, min_row=data_start, max_row=data_end)
        cats = Reference(ws_chart, min_col=1, min_row=data_start + 1, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws_chart.add_chart(chart, f"E{start_row + 1}")
        start_row = data_end + 8

    ws_main = wb_kpi.active
    if ws_main.max_row > 1:
        table = Table(displayName="KPI_Table", ref=f"A1:F{ws_main.max_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        table.tableStyleInfo = style
        ws_main.add_table(table)

    wb_kpi.save(kpi_file)

    # === Export PDFs ===
    print("\n📤 Exporting PDFs...")
    pdf_root = os.path.join(base_path, "Mapas_Gerados", "PDF", f"{mes}_{ano}")
    os.makedirs(pdf_root, exist_ok=True)

    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    for file in os.listdir(output_folder):
        if file.endswith(".xlsx") and not file.startswith("~$"):
            company = file.replace(f"{mes}_", "").replace(".xlsx", "")
            company_folder = os.path.join(pdf_root, company)
            os.makedirs(company_folder, exist_ok=True)

            wb = excel.Workbooks.Open(os.path.join(output_folder, file))
            for sheet in wb.Sheets:
                employee = sheet.Name.strip().replace(" ", "_")
                pdf_path = os.path.join(company_folder, f"{employee}.pdf")

                sheet.PageSetup.Orientation = 2  # Landscape
                sheet.PageSetup.Zoom = False
                sheet.PageSetup.FitToPagesTall = 1
                sheet.PageSetup.FitToPagesWide = 1
                sheet.PageSetup.CenterHorizontally = True

                # Optional: adjust margins
                sheet.PageSetup.LeftMargin = sheet.PageSetup.Application.InchesToPoints(0.3)
                sheet.PageSetup.RightMargin = sheet.PageSetup.Application.InchesToPoints(0.3)
                sheet.PageSetup.TopMargin = sheet.PageSetup.Application.InchesToPoints(0.5)
                sheet.PageSetup.BottomMargin = sheet.PageSetup.Application.InchesToPoints(0.5)

                sheet.ExportAsFixedFormat(0, pdf_path)

            wb.Close(SaveChanges=False)

    excel.Quit()
    print(f"✅ PDFs exported to: {pdf_root}")
